"""
建材直供/出库量 增量更新工具
============================
功能：
    把"建筑钢材出库量"Mysteel Excel（按日期倒序）按日期升序追加到
    "钢材直供量统计.xlsx"的"源数据-更新"sheet 的 A-M 列，
    并按现有公式模式自动补全 N-AH 列。

公式模式（已从目标表实际样例提取）：
    N..Y  (col 14..25)  MA5   =AVERAGE({col}{r-4}:{col}{r})
    Z..AC (col 26..29)  拿货  =线盘MA5 + 螺纹MA5 - 直供MA5
    AD..AG(col 30..33)  占比  =直供MA5 / (线盘MA5 + 螺纹MA5)
    AH    (col 34)      差值  =J{r} - B{r}

命令行：
    python update_zhigong.py <源Excel路径>
"""

from __future__ import annotations

import argparse
import datetime as dt
import shutil
import sys
from pathlib import Path

import openpyxl

# ============= 默认路径 =============
DEFAULT_SRC_PATTERN = "1.1全国建筑钢材钢企日出库量监测（三大区域）"
DEFAULT_TGT_PATH = Path(
    r"C:\Users\Administrator\Nutstore\1\小目标\钢材直供量统计.xlsx"
)
DEFAULT_TGT_SHEET = "源数据-更新"
# 目标表头：列名->列号（1-based），与原表保持一致
COL_NAMES = [
    "日期",
    "建材直供量-全国", "建材直供量-华东", "建材直供量-南方", "建材直供量-北方",
    "线盘出库量-全国", "线盘出库量-华东", "线盘出库量-南方", "线盘出库量-北方",
    "螺纹出库量-全国", "螺纹出库量-华东", "螺纹出库量-南方", "螺纹出库量-北方",
]
# 13 列数据，对应目标表 A..M


# ============= 工具函数 =============
def _to_date(v) -> dt.date:
    """统一把单元格值转成 date；支持 datetime/字符串'YYYY/M/D'。"""
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, str):
        v = v.strip()
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
            try:
                return dt.datetime.strptime(v, fmt).date()
            except ValueError:
                continue
    raise ValueError(f"无法解析日期: {v!r}")


def _find_src_sheet(wb, hint: str) -> str:
    """找到源表里第一个非目录 sheet。"""
    for name in wb.sheetnames:
        if name == "目录":
            continue
        return name
    return hint  # fallback


def load_source_records(src_path: Path) -> list[tuple[dt.date, list[float]]]:
    """读取源 Excel，返回 [(date, [B..M 12 个数值])] 升序。"""
    wb = openpyxl.load_workbook(src_path, data_only=True)
    ws = wb[_find_src_sheet(wb, DEFAULT_SRC_PATTERN)]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if d is None:
            continue
        try:
            date = _to_date(d)
        except ValueError:
            continue  # 跳过表尾"返回目录"等非数据行
        nums = list(row[1:13])
        if any(v is None for v in nums):
            continue
        records.append((date, [float(v) for v in nums]))
    records.sort(key=lambda x: x[0])  # 升序
    return records


def find_last_date_in_target(ws) -> dt.date | None:
    """找到目标 sheet 最后一行的日期。"""
    for r in range(ws.max_row, 1, -1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        try:
            return _to_date(v)
        except ValueError:
            continue
    return None


# ============= 公式生成 =============
def build_formulas(r: int) -> dict[str, str]:
    """对行 r 生成 N..AH 全部公式，返回 {坐标: 公式}。"""
    f: dict[str, str] = {}
    # N..Y (14..25) MA5，B..M 各自 5 日均线
    for col_idx, col_letter in enumerate("BCDEFGHIJKLM", start=2):
        target_col = openpyxl.utils.get_column_letter(col_idx + 12)  # +12 → N..
        f[f"{target_col}{r}"] = f"=AVERAGE({col_letter}{r-4}:{col_letter}{r})"
    # Z..AC (26..29) 贸易商拿货 = 线盘MA5 + 螺纹MA5 - 直供MA5
    zips = [("Z", "N", "R", "V"),
            ("AA", "O", "S", "W"),
            ("AB", "P", "T", "X"),
            ("AC", "Q", "U", "Y")]
    for tgt, zhigong, xianpan, luowen in zips:
        f[f"{tgt}{r}"] = f"={xianpan}{r}+{luowen}{r}-{zhigong}{r}"
    # AD..AG (30..33) 直供占比 = 直供MA5 / (线盘MA5 + 螺纹MA5)
    ratios = [("AD", "N", "R", "V"),
              ("AE", "O", "S", "W"),
              ("AF", "P", "T", "X"),
              ("AG", "Q", "U", "Y")]
    for tgt, zhigong, xianpan, luowen in ratios:
        f[f"{tgt}{r}"] = f"={zhigong}{r}/({xianpan}{r}+{luowen}{r})"
    # AH (34) 螺纹-直供
    f[f"AH{r}"] = f"=J{r}-B{r}"
    return f


# ============= 主流程 =============
def update(
    src_path: Path,
    tgt_path: Path = DEFAULT_TGT_PATH,
    tgt_sheet: str = DEFAULT_TGT_SHEET,
    dry_run: bool = False,
) -> dict:
    """追加新数据到目标 sheet。返回 {added:[(date, row)], skipped:[date]}。"""
    if not src_path.exists():
        raise FileNotFoundError(src_path)
    if not tgt_path.exists():
        raise FileNotFoundError(tgt_path)

    records = load_source_records(src_path)
    if not records:
        raise ValueError("源 Excel 没有可识别的数据行")

    wb = openpyxl.load_workbook(tgt_path)
    ws = wb[tgt_sheet]
    last_date = find_last_date_in_target(ws)
    if last_date is None:
        raise RuntimeError("目标 sheet 找不到已有日期，无法定位追加位置")

    # 备份
    backup = tgt_path.with_suffix(tgt_path.suffix + ".bak")
    if not dry_run:
        shutil.copy2(tgt_path, backup)

    # 找最后连续行号
    last_row = ws.max_row
    while last_row > 0 and ws.cell(last_row, 1).value is None:
        last_row -= 1

    new_records = [r for r in records if r[0] > last_date]
    added: list[tuple[dt.date, int]] = []
    skipped: list[dt.date] = []

    # 保留"目标已有最后日期"那行的 N..AH 公式不动（避免破坏图表源数据）
    # 从 last_row + 1 开始写新行
    row = last_row + 1
    for date, values in new_records:
        # A 列日期
        ws.cell(row, 1, date)
        # B..M 数据
        for i, v in enumerate(values, start=2):
            ws.cell(row, i, v)
        # N..AH 公式
        if not dry_run:
            for coord, formula in build_formulas(row).items():
                # coord 形如 "N2223"，拆成 (col_letter, row) 喂给 cell
                col_letter = ''.join(ch for ch in coord if ch.isalpha())
                ws.cell(row, openpyxl.utils.column_index_from_string(col_letter), formula)
        added.append((date, row))
        row += 1

    # 报告
    if not dry_run:
        wb.save(tgt_path)

    return {
        "last_date_before": last_date,
        "backup": str(backup) if not dry_run else None,
        "added": added,
    }


# ============= CLI =============
def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("src", nargs="?", help="源 Excel 路径（建筑钢材出库量）")
    ap.add_argument("--tgt", default=str(DEFAULT_TGT_PATH), help="目标 Excel 路径")
    ap.add_argument("--sheet", default=DEFAULT_TGT_SHEET, help="目标 sheet 名")
    ap.add_argument("--dry-run", action="store_true", help="只读不写")
    args = ap.parse_args()

    src = Path(args.src) if args.src else None
    if src is None:
        print("用法: python update_zhigong.py <源Excel路径> [--tgt 目标] [--sheet sheet名]")
        sys.exit(1)

    result = update(Path(src), Path(args.tgt), args.sheet, args.dry_run)
    print(f"目标表最后已有日期: {result['last_date_before']}")
    if result["backup"]:
        print(f"已备份: {result['backup']}")
    print(f"本次追加 {len(result['added'])} 行：")
    for d, r in result["added"]:
        print(f"  row {r}: {d}")
    print("完成。")


if __name__ == "__main__":
    main()
