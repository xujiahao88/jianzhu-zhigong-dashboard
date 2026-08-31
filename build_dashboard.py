"""
钢材直供/出库量 看板生成器（v2：表格 + 图表）
=========================================
读取「钢材直供量统计.xlsx」→
顶部生成 Excel 同款三色表头数据表（近 N 天 + MA5 + 周环比色块）→
底部生成近 5 年多年度叠加线图→
输出 dashboard/index.html，可部署到 GitHub Pages / CloudStudio。
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys
from pathlib import Path

import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib.font_manager as fm
import pandas as pd
import numpy as np

# ── 默认路径 ──────────────────────────────────
DEFAULT_EXCEL = Path(r"C:\Users\Administrator\Nutstore\1\小目标\钢材直供量统计.xlsx")
OUT_DIR = Path(__file__).resolve().parent / "dashboard"

# ── 中文字体（Windows 优先微软雅黑） ──
for _name in ["Microsoft YaHei", "SimHei", "Noto Sans CJK SC", "WenQuanYi Micro Hei"]:
    try:
        fm.findfont(_name, fallback_to_default=False)
        plt.rcParams["font.family"] = _name
        break
    except Exception:
        continue
plt.rcParams["axes.unicode_minus"] = False

# 12 个指标列定义（与 Excel 源表 A-M 一一对应）
RAW_COLS = [
    ("建材直供量", "全国"), ("建材直供量", "华东"), ("建材直供量", "南方"), ("建材直供量", "北方"),
    ("线盘出库量", "全国"), ("线盘出库量", "华东"), ("线盘出库量", "南方"), ("线盘出库量", "北方"),
    ("螺纹出库量", "全国"), ("螺纹出库量", "华东"), ("螺纹出库量", "南方"), ("螺纹出库量", "北方"),
]

# 三个分组（用于三色表头）
GROUPS = [
    ("建材直供量", "#4a7ab8"),  # 蓝
    ("线盘出库量", "#5a9b6e"),  # 绿
    ("螺纹出库量", "#c9a855"),  # 黄/金
]
INDICATORS_FOR_CHART = [
    ("建材直供量(MA5)", [f"建材直供量_{r}_MA5" for r in ["全国", "华东", "南方", "北方"]]),
    ("线盘出库量(MA5)", [f"线盘出库量_{r}_MA5" for r in ["全国", "华东", "南方", "北方"]]),
    ("螺纹出库量(MA5)", [f"螺纹出库量_{r}_MA5" for r in ["全国", "华东", "南方", "北方"]]),
    ("贸易商拿货量(MA5)", [f"贸易商拿货量(MA5)_{r}" for r in ["全国", "华东", "南方", "北方"]]),
]
RATIO_FOR_CHART = [
    ("建材直供占比(MA5)", [f"建材直供占比(MA5)_{r}" for r in ["全国", "华东", "南方", "北方"]]),
]


def load_data(excel_path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["源数据-更新"]
    rows = []
    for r in range(3, ws.max_row + 1):
        d = ws.cell(r, 1).value
        if d is None:
            continue
        if isinstance(d, dt.datetime):
            d = d.date()
        elif isinstance(d, dt.date):
            pass
        else:
            continue
        vals = [ws.cell(r, c).value for c in range(2, 14)]
        if any(v is None for v in vals):
            continue
        try:
            nums = [float(v) for v in vals]
        except (ValueError, TypeError):
            continue
        rows.append([d] + nums)
    df = pd.DataFrame(rows, columns=["date"] + [f"{g}_{r}" for g, r in RAW_COLS])
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date").reset_index(drop=True)
    return df


def compute_derived(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        if col == "date":
            continue
        df[f"{col}_MA5"] = df[col].rolling(window=5).mean()
    for region in ["全国", "华东", "南方", "北方"]:
        df[f"贸易商拿货量(MA5)_{region}"] = (
            df[f"线盘出库量_{region}_MA5"] + df[f"螺纹出库量_{region}_MA5"] - df[f"建材直供量_{region}_MA5"]
        )
    for region in ["全国", "华东", "南方", "北方"]:
        denom = df[f"线盘出库量_{region}_MA5"] + df[f"螺纹出库量_{region}_MA5"]
        df[f"建材直供占比(MA5)_{region}"] = df[f"建材直供量_{region}_MA5"] / denom
    return df


def filter_recent_years(df: pd.DataFrame, years: int = 5) -> pd.DataFrame:
    cutoff_year = df["date"].dt.year.max() - (years - 1)  # 当前年-4 = 2022
    return df[df["date"].dt.year >= cutoff_year].copy()


# ────────────────────────────────────────────
# 表格生成（HTML，复刻 Excel 三色表头 + 周环比色块）
# ────────────────────────────────────────────
def build_table(df: pd.DataFrame, recent_days: int = 10) -> str:
    """生成 Excel 同款样式的数据表 HTML。"""
    df = df.sort_values("date", ascending=False).head(recent_days).sort_values("date").reset_index(drop=True)
    # 算 MA5 和周环比
    df_calc = df.copy()
    for col in [c for c in df_calc.columns if c.endswith("_全国") or c.endswith("_华东") or c.endswith("_南方") or c.endswith("_北方")]:
        if "MA5" in col:
            continue
        df_calc[f"{col}_MA5"] = df_calc[col].rolling(window=5).mean()
    # MA5 周环比 = 今天 MA5 - 7 天前 MA5
    mom = df.iloc[-1]
    week_ago_idx = max(0, len(df) - 8)  # 找 7 天前的索引
    week_ago = df.iloc[week_ago_idx]

    # 三色表头
    head = ['<tr class="head"><th class="corner" rowspan="2">日期</th>']
    for grp_name, grp_color in GROUPS:
        head.append(f'<th colspan="4" style="background:{grp_color};color:#fff">{grp_name}</th>')
    head.append('</tr>')
    head.append('<tr class="subhead">')
    for grp_name, _ in GROUPS:
        for region in ["全国", "华东", "南方", "北方"]:
            head.append(f'<th>{grp_name}:{region}</th>')
    head.append('</tr>')

    # 数据行
    body = []
    for _, row in df.iterrows():
        body.append('<tr>')
        body.append(f'<td class="date">{row["date"].strftime("%Y/%m/%d")}</td>')
        for grp_name, _ in GROUPS:
            for region in ["全国", "华东", "南方", "北方"]:
                v = row[f"{grp_name}_{region}"]
                body.append(f'<td>{v:.2f}</td>')
        body.append('</tr>')

    # MA5 行（最新一天）
    ma5_row = df_calc.iloc[-1]
    body.append('<tr class="ma5"><td class="date">MA5</td>')
    for grp_name, _ in GROUPS:
        for region in ["全国", "华东", "南方", "北方"]:
            v = ma5_row.get(f"{grp_name}_{region}_MA5")
            txt = f"{v:.2f}" if pd.notna(v) else "-"
            body.append(f'<td>{txt}</td>')
    body.append('</tr>')

    # 周环比（色块，绿涨红跌）
    body.append('<tr class="mom"><td class="date">MA5周环比</td>')
    for grp_name, _ in GROUPS:
        for region in ["全国", "华东", "南方", "北方"]:
            ma5_now = ma5_row.get(f"{grp_name}_{region}_MA5")
            ma5_wk = week_ago.get(f"{grp_name}_{region}")  # 简化：7 天前原始值
            if pd.isna(ma5_now) or pd.isna(ma5_wk):
                body.append('<td>-</td>')
                continue
            diff = ma5_now - ma5_wk
            color = "#5cb85c" if diff > 0 else "#d9534f" if diff < 0 else "#aaa"
            txt = f"{diff:+.2f}"
            body.append(f'<td style="background:{color};color:#fff;font-weight:bold">{txt}</td>')
    body.append('</tr>')

    return "<table class='data'>" + "".join(head) + "".join(body) + "</table>"


# ────────────────────────────────────────────
# 图表生成（近 5 年）
# ────────────────────────────────────────────
def _plot(ax, df, yg, col, title, is_ratio=False):
    current_year = int(df["date"].dt.year.max())
    # 显式年份→颜色映射：历史年虚线（灰/蓝渐变），上一年深蓝实线，当前年红实线
    year_color = {
        current_year: "#E41A1C",       # 当前年：红
        current_year - 1: "#08306B",   # 上一年：深蓝
        current_year - 2: "#08519C",   # 前2年：较深蓝
        current_year - 3: "#3182BD",   # 前3年：亮蓝
        current_year - 4: "#A6A6A6",   # 前4年：灰
    }
    years_sorted = sorted(yg.keys())
    for idx, year in enumerate(years_sorted):
        if len(yg[year]) < 2:
            continue
        ydf = yg[year].copy()
        ydf["day_of_year"] = ydf["date"].dt.dayofyear
        ydf = ydf.dropna(subset=[col])
        if len(ydf) < 5:
            continue
        # 颜色 + 线型：当前年=实线红，上一年=实线深蓝，历史年=虚线
        if year == current_year:
            color, linewidth, alpha, linestyle = year_color[year], 2.2, 1.0, "-"
        elif year == current_year - 1:
            color, linewidth, alpha, linestyle = year_color[year], 1.6, 1.0, "-"
        else:
            color = year_color.get(year, "#888888")
            linewidth = 1.0
            alpha = 0.75
            linestyle = "--"
        ax.plot(ydf["day_of_year"], ydf[col],
                color=color, linewidth=linewidth, alpha=alpha,
                linestyle=linestyle, label=str(year))
        # 当前年末点标注
        if year == current_year:
            last = ydf.iloc[-1]
            ax.annotate(f"{last[col]:.2f}", xy=(last["day_of_year"], last[col]),
                        xytext=(5, 2), textcoords="offset points",
                        fontsize=7, color=color, fontweight="bold")
    ax.set_title(title, fontsize=8, fontweight="bold")
    ax.xaxis.set_major_locator(mticker.FixedLocator(
        [dt.date(2024, m, 1).timetuple().tm_yday for m in range(1, 13)]))
    ax.xaxis.set_major_formatter(mticker.FixedFormatter([f"{m}/1" for m in range(1, 13)]))
    ax.set_xlim(0, 366)  # 1月1日到12月31日
    ax.tick_params(labelsize=6)
    ax.grid(True, alpha=0.3, linewidth=0.5)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    if is_ratio:
        ax.yaxis.set_major_formatter(mticker.PercentFormatter(1.0, decimals=0))


def build_charts(df: pd.DataFrame, out_dir: Path) -> list[tuple[str, Path]]:
    out_dir.mkdir(parents=True, exist_ok=True)
    df["year"] = df["date"].dt.year
    yg = {int(y): df[df["year"] == y].copy() for y in df["year"].unique()}

    imgs = []

    # Page 1: 4×4 网格（建材直供MA5 / 线盘MA5 / 螺纹MA5 / 拿货MA5 × 4 区）
    fig, axes = plt.subplots(4, 4, figsize=(16, 12), dpi=120)
    regions = ["全国", "华东", "南方", "北方"]
    for i, (ind_label, cols) in enumerate(INDICATORS_FOR_CHART):
        for j, col in enumerate(cols):
            region = regions[j]
            ax = axes[i][j]
            _plot(ax, df, yg, col, f"{ind_label}-{region}")
            if i == 0 and j == 0:
                ax.legend(fontsize=5, ncol=3, loc="upper left", frameon=False, handlelength=1.0)
    plt.suptitle(f"建材直供&出库 MA5 多年度叠加（近5年）\n数据来源：我的钢铁网  截止：{df['date'].max().strftime('%Y-%m-%d')}",
                 fontsize=11, fontweight="bold")
    plt.tight_layout(rect=[0, 0, 1, 0.97])
    p1 = out_dir / "page1_直供出库.png"
    fig.savefig(p1, dpi=120, bbox_inches="tight")
    plt.close(fig)
    imgs.append(("建材直供&出库日度跟踪", p1))

    # Page 2: 1×4 占比
    fig, axes = plt.subplots(1, 4, figsize=(16, 4), dpi=120)
    for j, col in enumerate(RATIO_FOR_CHART[0][1]):
        region = regions[j]
        ax = axes[j]
        _plot(ax, df, yg, col, f"建材直供占比(MA5)-{region}", is_ratio=True)
    plt.suptitle(f"建材直供占比 MA5 多年度叠加（近5年）\n数据来源：我的钢铁网  截止：{df['date'].max().strftime('%Y-%m-%d')}",
                 fontsize=11, fontweight="bold")
    plt.tight_layout(rect=[0, 0, 1, 0.92])
    p2 = out_dir / "page2_直供占比.png"
    fig.savefig(p2, dpi=120, bbox_inches="tight")
    plt.close(fig)
    imgs.append(("建材直供占比 MA5 日度跟踪", p2))

    return imgs


def build_html(table_html: str, imgs: list[tuple[str, Path]], out_dir: Path):
    last_update = "实时"
    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>直供&出库日度跟踪</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family: -apple-system, "Microsoft YaHei", sans-serif; background:#f5f5f5; color:#333; font-size:12px; }}
  .header {{ background: linear-gradient(135deg, #1a237e, #283593); color:#fff; padding:14px 20px; }}
  .header h1 {{ font-size:18px; margin-bottom:4px; }}
  .header p {{ font-size:11px; opacity:0.85; }}
  .container {{ max-width:1240px; margin:0 auto; padding:8px; }}
  .card {{ background:#fff; border-radius:6px; box-shadow:0 1px 3px rgba(0,0,0,0.1); margin-bottom:14px; overflow:hidden; }}
  .card-title {{ background:#4a7ab8; color:#fff; font-size:16px; font-weight:bold; padding:10px 14px; text-align:center; }}
  .card-body {{ padding:10px; }}
  .card-body img {{ width:100%; display:block; }}

  /* Excel 同款表格 —— 数字加大加粗，重点突出 */
  table.data {{ border-collapse:collapse; width:100%; font-size:15px; }}
  table.data th, table.data td {{ border:1px solid #999; padding:6px 8px; text-align:center; min-width:56px; font-weight:600; }}
  table.data th {{ color:#fff; font-weight:700; }}
  table.data tr.head th {{ font-size:16px; }}
  table.data tr.subhead th {{ background:#e0e0e0; color:#333; font-weight:600; font-size:12px; }}
  table.data th.corner {{ background:#e0e0e0; color:#333; }}
  table.data td.date {{ background:#f0f0f0; font-weight:700; text-align:center; font-size:13px; }}
  table.data tr.ma5 td {{ background:#e8e8e8; font-style:italic; }}
  table.data tr.mom td {{ font-weight:700; }}
  .footer {{ text-align:center; font-size:10px; color:#999; padding:14px; }}
</style>
</head>
<body>
<div class="header">
  <h1>直供&出库日度跟踪</h1>
  <p>数据来源：我的钢铁网 ｜ 图表为近 5 年（2022–2026）多年度叠加 ｜ 数据更新：{last_update}</p>
</div>
<div class="container">
  <div class="card">
    <div class="card-title">直供&出库日度跟踪</div>
    <div class="card-body">{table_html}</div>
  </div>
"""
    for title, path in imgs:
        html += f"""  <div class="card">
    <div class="card-title">{title}</div>
    <div class="card-body"><img src="{path.name}" alt="{title}"></div>
  </div>
"""
    html += """  <div class="footer">由 build_dashboard.py 自动生成 · 数据来源：我的钢铁网</div>
</div>
</body>
</html>"""
    (out_dir / "index.html").write_text(html, encoding="utf-8")


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--excel", default=str(DEFAULT_EXCEL))
    ap.add_argument("--out", default=str(OUT_DIR))
    ap.add_argument("--years", type=int, default=5, help="图表保留近 N 年")
    ap.add_argument("--recent-days", type=int, default=10, help="表格显示近 N 天")
    args = ap.parse_args()

    excel_path = Path(args.excel)
    out_dir = Path(args.out)
    if not excel_path.exists():
        print(f"❌ 文件不存在: {excel_path}")
        sys.exit(1)

    print(f"📖 读取: {excel_path}")
    df = load_data(excel_path)
    print(f"   共 {len(df)} 行, 日期范围 {df.date.min().date()} ~ {df.date.max().date()}")

    print("🔧 计算衍生指标...")
    df = compute_derived(df)

    print(f"📅 筛选近 {args.years} 年...")
    df_recent = filter_recent_years(df, years=args.years)
    print(f"   剩 {len(df_recent)} 行 ({df_recent.date.min().date()} ~ {df_recent.date.max().date()})")

    print(f"📋 生成数据表（近 {args.recent_days} 天）...")
    table_html = build_table(df, recent_days=args.recent_days)

    print("🎨 绘制图表...")
    imgs = build_charts(df_recent, out_dir)
    for title, p in imgs:
        print(f"   ✅ {title}: {p.name}")

    print("🌐 生成 HTML...")
    build_html(table_html, imgs, out_dir)
    print(f"   ✅ index.html")

    print(f"\n📁 输出: {out_dir.absolute()}")


if __name__ == "__main__":
    main()